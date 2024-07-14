from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
import pandas as pd
import openpyxl
import io
import traceback

app = FastAPI()

# Allow CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/uploadfile/")
async def upload_file(file: UploadFile = File(...)):
    try:
        # Leer el archivo Excel usando openpyxl
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = wb.active

        # Verificación básica del DataFrame
        if sheet.max_row == 0 or sheet.max_column == 0:
            raise HTTPException(status_code=400, detail="El archivo Excel está vacío o tiene un formato incorrecto.")
        
        # Procesar el contenido del archivo
        tipo_hora = sheet.cell(row=1, column=3).value
        result = {'maquina': {}}

        current_machine = None
        machine_time_day = []
        machine_hours_day = []
        stop_time_day = []
        machine_time_night = []
        machine_hours_night = []
        stop_time_nigth = []
        leftover = {}
        tiempo_total = None
        cantidad_total = None

        for row in sheet.iter_rows(min_row=2, max_col=12, values_only=False):
            cell = row[2]
            value = cell.value
            if cell.font.bold:  # Detectar nombres de máquinas en negrita
                # Agregar un diccionario dependiendo de los turnos
                if current_machine and leftover:
                    result['maquina'][current_machine] = {
                        'tiempo_total': tiempo_total,
                        'cantidad_total': cantidad_total,
                        'informe_detallado': leftover
                    }
                current_machine = value
                tiempo_total = row[8].value
                cantidad_total = row[11].value

                if 'diurno' in leftover and machine_time_day:
                    leftover['diurno'] = {
                        'tiempo_ejecucion': machine_time_day,
                        'tiempo_maquina': machine_hours_day,
                        'tiempo_paro': stop_time_day,
                    }
                if 'nocturno' in leftover and machine_time_night:
                    leftover['nocturno'] = {
                        'tiempo_ejecucion': machine_time_night,
                        'tiempo_maquina': machine_hours_night,
                        'tiempo_paro': stop_time_nigth,
                    }

                # Ultimo registro de la máquina
                leftover = {}
                machine_time_day = []
                machine_hours_day = []
                stop_time_day = []
                machine_time_night = []
                machine_hours_night = []
                stop_time_nigth = []
            else:
                turno = row[6].value
                # Agregar un diccionario dependiendo de los turnos
                if turno not in leftover:
                    if turno == 1:
                        leftover['diurno'] = {}
                    elif turno == 2:
                        leftover['nocturno'] = {}
                
                if turno == 1:
                    # Agregar los tiempos a los diccionarios
                    if 'Horas ejecucion' in value:
                        machine_time_day.append({
                            'tiempo': row[8].value,
                            'cantidad': row[11].value,
                        })
                    elif 'Horas maquina' in value:
                        machine_hours_day.append({
                            'tiempo': row[8].value,
                            'cantidad': row[11].value,
                        })
                    elif 'Tiempo paro' in value:
                        stop_time_day.append({
                            'tiempo': row[8].value,
                            'cantidad': row[11].value,
                        })
                elif turno == 2:
                    # Agregar los tiempos a los diccionarios
                    if 'Horas ejecucion' in value:
                        machine_time_night.append({
                            'tiempo': row[8].value,
                            'cantidad': row[11].value,
                        })
                    elif 'Horas maquina' in value:
                        machine_hours_night.append({
                            'tiempo': row[8].value,
                            'cantidad': row[11].value,
                        })
                    elif 'Tiempo paro' in value:
                        stop_time_nigth.append({
                            'tiempo': row[8].value,
                            'cantidad': row[11].value,
                        })

        # Agregar la última máquina y sus tiempos
        if current_machine and leftover:
            if 'diurno' in leftover and machine_time_day:
                leftover['diurno'] = {
                    'tiempo_ejecucion': machine_time_day,
                    'tiempo_maquina': machine_hours_day,
                    'tiempo_paro': stop_time_day,
                }
            if 'nocturno' in leftover and machine_time_night:
                leftover['nocturno'] = {
                    'tiempo_ejecucion': machine_time_night,
                    'tiempo_maquina': machine_hours_night,
                    'tiempo_paro': stop_time_nigth,
                }
            result['maquina'][current_machine] = {
                'tiempo_total': tiempo_total,
                'cantidad_total': cantidad_total,
                'informe_detallado': leftover
                }
        return JSONResponse(content=result)
    except Exception as e:
        print(str(e))
        print(traceback.extract_tb(e.__traceback__))
        raise HTTPException(status_code=400, detail=f"Error procesando el archivo Excel: {e}")

@app.get("/")
async def main():
    return FileResponse("static/index.html")
